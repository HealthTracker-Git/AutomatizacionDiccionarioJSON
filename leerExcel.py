#%%
from openpyxl import load_workbook
import pandas as pd

def leer_archivo_xlsx(file_path):
    """
    Lee un archivo .xlsx y devuelve el objeto de la hoja activa para trabajar con sus datos.
    
    Args:
        file_path (str): Ruta del archivo .xlsx a leer.
    
    Returns:
        sheet (openpyxl.worksheet.worksheet.Worksheet): La hoja activa del archivo .xlsx.
    """
    # Cargar el libro de trabajo
    workbook = load_workbook(file_path)
    
    # Obtener la hoja activa (la primera hoja por defecto)
    sheet = workbook["A01"]
    
    
    # Imprimir el nombre de la hoja activa
    print(f"Hoja activa: {sheet.title}")
    
    #iterate cells, set value of cells with no left border to None
    for row in sheet.iter_rows():
        for cell in row:
            if not cell.border.left.style:
                try:
                    cell.value = None
                except:
                    continue

    df = pd.DataFrame(sheet.values).dropna(how='all', axis=0).dropna(how='all', axis=1).reset_index(drop=True) # load into pandas, drop empty rows and columns
    df = df.T.set_index(0).T # set first row as header

# Ejemplo de uso
file_path = 'DICCIONARIO_SERIE_A_2009.xlsx'
sheet = leer_archivo_xlsx(file_path)

# %% stackoverflow
def parse_excel_sheet(file, sheet_name=0, threshold=5):
    '''parses multiple tables from an excel sheet into multiple data frame objects. Returns [dfs, df_mds], where dfs is a list of data frames and df_mds their potential associated metadata'''
    xl = pd.ExcelFile(file)
    entire_sheet = xl.parse(sheet_name=sheet_name)

    # count the number of non-Nan cells in each row and then the change in that number between adjacent rows
    n_values = np.logical_not(entire_sheet.isnull()).sum(axis=1)
    n_values_deltas = n_values[1:] - n_values[:-1].values

    # define the beginnings and ends of tables using delta in n_values
    table_beginnings = n_values_deltas > threshold
    table_beginnings = table_beginnings[table_beginnings].index
    table_endings = n_values_deltas < -threshold
    table_endings = table_endings[table_endings].index
    if len(table_beginnings) < len(table_endings) or len(table_beginnings) > len(table_endings)+1:
        raise BaseException('Could not detect equal number of beginnings and ends')

    # look for metadata before the beginnings of tables
    md_beginnings = []
    for start in table_beginnings:
        md_start = n_values.iloc[:start][n_values==0].index[-1] + 1
        md_beginnings.append(md_start)

    # make data frames
    dfs = []
    df_mds = []
    for ind in range(len(table_beginnings)):
        start = table_beginnings[ind]+1
        if ind < len(table_endings):
            stop = table_endings[ind]
        else:
            stop = entire_sheet.shape[0]
        df = xl.parse(sheet_name=sheet_name, skiprows=start, nrows=stop-start)
        dfs.append(df)

        md = xl.parse(sheet_name=sheet_name, skiprows=md_beginnings[ind], nrows=start-md_beginnings[ind]-1).dropna(axis=1)
        df_mds.append(md)
    return dfs, df_mds
